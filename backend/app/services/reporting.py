import os
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from sqlalchemy.orm import Session
from ..database.models import Customer, Loan, Prediction

from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[2]
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# ----------------- CSV GENERATION -----------------
def generate_portfolio_csv(db: Session) -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Loan ID", "Customer Name", "Email", "Loan Amount ($)", "Term (Months)", 
        "Credit Score", "DTI Ratio", "Existing Loans", "Purpose", "Status", "Created At"
    ])
    
    # Query data
    loans = db.query(Loan).join(Customer).all()
    for loan in loans:
        writer.writerow([
            loan.id,
            f"{loan.customer.first_name} {loan.customer.last_name}",
            loan.customer.email,
            loan.loan_amount,
            loan.loan_term,
            loan.credit_score,
            loan.debt_to_income_ratio,
            loan.existing_loans,
            loan.loan_purpose,
            loan.status,
            loan.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    bytes_io = io.BytesIO()
    bytes_io.write(output.getvalue().encode('utf-8'))
    bytes_io.seek(0)
    return bytes_io


# ----------------- EXCEL GENERATION -----------------
def generate_portfolio_excel(db: Session) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Portfolio"
    
    # Title block
    ws.merge_cells('A1:K1')
    ws['A1'] = "Loan Default Prediction System - Portfolio Report"
    ws['A1'].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Metadata block
    ws['A3'] = f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(italic=True)
    
    # Headers
    headers = [
        "Loan ID", "Customer Name", "Email", "Loan Amount ($)", "Term (Months)", 
        "Credit Score", "DTI Ratio", "Existing Loans", "Purpose", "Status", "Created At"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        
    ws.row_dimensions[5].height = 25
    
    # Data Rows
    loans = db.query(Loan).join(Customer).all()
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    row_idx = 6
    for loan in loans:
        ws.cell(row=row_idx, column=1, value=loan.id).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=f"{loan.customer.first_name} {loan.customer.last_name}")
        ws.cell(row=row_idx, column=3, value=loan.customer.email)
        
        amount_cell = ws.cell(row=row_idx, column=4, value=loan.loan_amount)
        amount_cell.number_format = "$#,##0.00"
        
        ws.cell(row=row_idx, column=5, value=loan.loan_term).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=loan.credit_score).alignment = Alignment(horizontal="center")
        
        dti_cell = ws.cell(row=row_idx, column=7, value=loan.debt_to_income_ratio)
        dti_cell.number_format = "0.00%"
        
        ws.cell(row=row_idx, column=8, value=loan.existing_loans).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=9, value=loan.loan_purpose)
        
        status_cell = ws.cell(row=row_idx, column=10, value=loan.status)
        status_cell.alignment = Alignment(horizontal="center")
        # Highlight status
        if loan.status == "Approved":
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # light green
        elif loan.status == "Defaulted" or loan.status == "Rejected":
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
            
        ws.cell(row=row_idx, column=11, value=loan.created_at.strftime("%Y-%m-%d %H:%M"))
        
        # Apply borders to all columns in this row
        for c in range(1, 12):
            ws.cell(row=row_idx, column=c).border = thin_border
            
        row_idx += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    bytes_io = io.BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)
    return bytes_io


# ----------------- PDF GENERATION -----------------
def generate_prediction_pdf(prediction_id: int, db: Session) -> io.BytesIO:
    pred = db.query(Prediction).filter(Prediction.id == prediction_id).first()
    if not pred:
        raise ValueError("Prediction not found")
        
    customer = pred.customer
    loan = pred.loan
    
    import json
    explanation = json.loads(pred.explanation_json)
    
    # Set up pdf stream
    bytes_io = io.BytesIO()
    doc = SimpleDocTemplate(
        bytes_io, 
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=45, bottomMargin=45
    )
    
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#2563EB'),
        spaceBefore=12,
        spaceAfter=8,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#374151')
    )
    
    bold_body_style = ParagraphStyle(
        'BoldBody',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    # Story flowables list
    story = []
    
    # 1. Header Title
    story.append(Paragraph("LOAN DEFAULT RISK ASSESSMENT REPORT", title_style))
    story.append(Paragraph(f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
    story.append(Spacer(1, 15))
    
    # 2. Risk Evaluation Highlight Box
    # Set up color coding
    risk_color = '#10B981'  # Green
    if pred.risk_rating == "Medium":
        risk_color = '#F59E0B'  # Yellow/Orange
    elif pred.risk_rating == "High":
        risk_color = '#EF4444'  # Red
        
    risk_summary_data = [
        [
            Paragraph("<b>Risk Classification:</b>", body_style), 
            Paragraph(f"<font color='{risk_color}'><b>{pred.risk_rating.upper()} RISK</b></font>", bold_body_style)
        ],
        [
            Paragraph("<b>Probability of Default:</b>", body_style), 
            Paragraph(f"<b>{pred.risk_probability * 100:.1f}%</b>", bold_body_style)
        ],
        [
            Paragraph("<b>Assessment Confidence:</b>", body_style), 
            Paragraph(f"<b>{pred.confidence_score * 100:.1f}%</b>", bold_body_style)
        ],
        [
            Paragraph("<b>Loan Eligibility Score:</b>", body_style), 
            Paragraph(f"<b>{explanation.get('eligibility_score', 0)} / 100</b>", bold_body_style)
        ]
    ]
    
    risk_table = Table(risk_summary_data, colWidths=[180, 320])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(risk_table)
    story.append(Spacer(1, 15))
    
    # 3. Customer Demographics & Financial Profile Table
    story.append(Paragraph("Customer Demographic & Financial Profile", h2_style))
    profile_data = [
        [
            Paragraph("<b>Customer Name:</b>", body_style), 
            Paragraph(f"{customer.first_name} {customer.last_name}", body_style),
            Paragraph("<b>Age / Gender:</b>", body_style), 
            Paragraph(f"{customer.age} / {customer.gender}", body_style)
        ],
        [
            Paragraph("<b>Annual Income:</b>", body_style), 
            Paragraph(f"${customer.annual_income:,.2f}", body_style),
            Paragraph("<b>Employment Type:</b>", body_style), 
            Paragraph(f"{customer.employment_type} ({customer.job_experience} yrs)", body_style)
        ],
        [
            Paragraph("<b>Education:</b>", body_style), 
            Paragraph(customer.education, body_style),
            Paragraph("<b>Marital Status:</b>", body_style), 
            Paragraph(customer.marital_status, body_style)
        ],
        [
            Paragraph("<b>Home Ownership:</b>", body_style), 
            Paragraph(customer.home_ownership, body_style),
            Paragraph("<b>Email / Phone:</b>", body_style), 
            Paragraph(f"{customer.email} / {customer.phone}", body_style)
        ]
    ]
    
    profile_table = Table(profile_data, colWidths=[110, 140, 110, 140])
    profile_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F1F5F9')),
    ]))
    story.append(profile_table)
    story.append(Spacer(1, 15))
    
    # 4. Loan Application details
    story.append(Paragraph("Loan Application & Credit Information", h2_style))
    loan_data = [
        [
            Paragraph("<b>Requested Amount:</b>", body_style),
            Paragraph(f"${loan.loan_amount:,.2f}", body_style),
            Paragraph("<b>Loan Term:</b>", body_style),
            Paragraph(f"{loan.loan_term} Months", body_style)
        ],
        [
            Paragraph("<b>Credit Score:</b>", body_style),
            Paragraph(str(loan.credit_score), body_style),
            Paragraph("<b>Debt-to-Income (DTI):</b>", body_style),
            Paragraph(f"{loan.debt_to_income_ratio * 100:.2f}%", body_style)
        ],
        [
            Paragraph("<b>Monthly EMI:</b>", body_style),
            Paragraph(f"${loan.emi:,.2f}", body_style),
            Paragraph("<b>Loan Purpose:</b>", body_style),
            Paragraph(loan.loan_purpose, body_style)
        ],
        [
            Paragraph("<b>Savings / Current Bal:</b>", body_style),
            Paragraph(f"${loan.savings_balance:,.0f} / ${loan.current_balance:,.0f}", body_style),
            Paragraph("<b>Previous Defaults:</b>", body_style),
            Paragraph(str(loan.previous_defaults), body_style)
        ]
    ]
    
    loan_table = Table(loan_data, colWidths=[110, 140, 110, 140])
    loan_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F1F5F9')),
    ]))
    story.append(loan_table)
    story.append(Spacer(1, 15))
    
    # 5. Model Explainability
    story.append(Paragraph("Model Interpretation (Top Risk Contributors)", h2_style))
    story.append(Paragraph("Below are the local feature attribution reasons identifying what elements shifted default probabilities compared to average values in our training dataset.", body_style))
    story.append(Spacer(1, 6))
    
    explain_headers = [["Risk Factor Details", "Attribution Shift"]]
    explain_data = []
    
    # Add Risk Increasers
    for item in explanation.get("risk_increasers", []):
        explain_data.append([
            Paragraph(f"🔴 {item['description']}", body_style),
            Paragraph(f"+{item['impact']*100:.1f}% risk increase", bold_body_style)
        ])
        
    # Add Risk Reducers
    for item in explanation.get("risk_reducers", []):
        explain_data.append([
            Paragraph(f"🟢 {item['description']}", body_style),
            Paragraph(f"{item['impact']*100:.1f}% risk reduction", bold_body_style)
        ])
        
    if not explain_data:
        explain_data.append([Paragraph("No major risk factors detected outside the median range.", body_style), "0.0%"])
        
    explain_table_data = explain_headers + explain_data
    explain_table = Table(explain_table_data, colWidths=[350, 150])
    explain_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    
    # Set text colors in table headers manually
    for idx, cell in enumerate(explain_table_data[0]):
        explain_table.setStyle(TableStyle([
            ('TEXTCOLOR', (idx, 0), (idx, 0), colors.white)
        ]))
    story.append(explain_table)
    
    # Build Document
    doc.build(story)
    bytes_io.seek(0)
    return bytes_io
